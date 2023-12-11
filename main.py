from PyQt5 import QtWidgets,QtCore,QtGui
from PyQt5.QtWidgets import QDialog
from PyQt5.QtGui import QDesktopServices,QMovie
from selenium.common.exceptions import WebDriverException
from urllib3.exceptions import MaxRetryError
from oracleUI import Ui_MainWindow
import openpyxl
import os
import datetime
from ActionsDialog import Actionswindow
from PyQt5.QtCore import QThread
from PyQt5.QtCore import pyqtSignal,Qt
from getupdatesContra1 import contractupdate,takedriverc,sender
from getupdatesPayCert2 import payrequpdate
from getupdatesVarReq3 import varrequpdate
from getupdatesVarOrdr4 import varordupdate
import time 
from logincredentials import Ui_Dialog
from goonline import onlinefunc


current_folder = os.getcwd()
date = datetime.date.today()

class goonlinethread(QThread):
    
    finished = pyqtSignal(bool,int)
    
    def __init__(self,hidechrome,email,passw):
        super().__init__()
        self.hidechrome = hidechrome
        self.email = email
        self.passw = passw
        
    def run(self):
        
        try:
            onlinefunc(self.hidechrome,self.email,self.passw)
            status = 1
        except WebDriverException:
            status = 0
        except MaxRetryError:
            status = 0
        except AttributeError:
            status = 0
        self.finished.emit(False,status)
        
class worker1(QThread):
    
    finished=pyqtSignal(bool,int)
    
    def __init__(self,hidechrome,email,passw):
        super().__init__()
        self.hidechrome = hidechrome
        self.email = email
        self.passw = passw
    
    def run(self):
        try:
            contractupdate(self.hidechrome,self.email,self.passw)
            status = 1
        except WebDriverException:
            status = 0
        except MaxRetryError:
            status = 0
        except AttributeError:
            status = 0
        self.finished.emit(False,status)
        
class worker2(QThread):
    
    finished=pyqtSignal(bool,int)
    
    def __init__(self,hidechrome,email,passw):
        super().__init__()
        self.hidechrome = hidechrome
        self.email = email
        self.passw = passw
        
    def run(self):
        try:
            payrequpdate(self.hidechrome,self.email,self.passw)
            status = 1
        except WebDriverException:
            status = 0
        except MaxRetryError:
            status = 0
        self.finished.emit(False,status)
        
class worker3(QThread):
    
    finished=pyqtSignal(bool,int)
    
    def __init__(self,hidechrome,email,passw):
        super().__init__()
        self.hidechrome = hidechrome
        self.email = email
        self.passw = passw
        
    def run(self):
        try:
            varrequpdate(self.hidechrome,self.email,self.passw)
            status = 1
        except WebDriverException:
            status = 0
        except MaxRetryError:
            status = 0
            
        self.finished.emit(False,status)
        
class worker4(QThread):
    
    finished=pyqtSignal(bool,int)
    
    def __init__(self,hidechrome,email,passw):
        super().__init__()
        self.hidechrome = hidechrome
        self.email = email
        self.passw = passw
        
    def run(self):
        try:
            varordupdate(self.hidechrome,self.email,self.passw)
            status = 1
        except WebDriverException:
            status = 0
        except MaxRetryError:
            status = 0   
        self.finished.emit(False,status)



class killer(QThread):
    finished= pyqtSignal(bool,int)
    def __init__(self,typee):
        super().__init__()
        self.typee = typee
    def run(self):
        try:
            if self.typee=='Contracts':
                takedriverc(kill=True)
            elif self.typee=='Payment Certificates':
                pass
            elif self.typee=='Variation Requests':
                pass
            elif self.typee=='Variation Orders':
                pass
            else:
                    print('There is something wrong with the selection')
        except NameError:
            pass
        self.finished.emit(True,10)
        
class mainwindow(Ui_MainWindow):

    def maindisp(self,window):
        super().setupUi(window)
        self.readandset(typee='Contracts')
        self.s = sender()
        self.myedit()
        container.setWindowIcon(QtGui.QIcon('logo.jpg'))
        container.setFixedSize(968, 765)
        
        self.s.status_signal.connect(self.updatestatus)
        
    def myedit(self):
        self.textfile = open('textboxesdata.txt','w',encoding='utf-8')
        self.textfile.write('')   
        self.openexcelbutton.clicked.connect(self.openexcel)
        self.openexcelbutton_2.clicked.connect(self.actionsWindow)
        self.comboBox.currentIndexChanged.connect(self.changeaspect)
        self.comboBox.currentIndexChanged.connect(self.readandset)
        self.closebutton.clicked.connect(container.close)
        self.updatestautsbutton.clicked.connect(self.getupdate)
        self.abortupdatebut.clicked.connect(self.abort)
        self.actionEdit_Log_In_Credentials.triggered.connect(self.loginedit)
        self.actionUpdate_Status.triggered.connect(self.getupdate)
        self.actionOpen_Excel_Sheet.triggered.connect(self.openexcel)
        self.checkBox.setChecked(False)
        self.abortupdatebut.hide()
        self.movie = QMovie('LOADING.GIF')
        self.loading.setMovie(self.movie)
        self.checkBox.setChecked(True)   
        self.updateall.setChecked(True)
        self.updateall.hide()
        self.timer.hide()
        self.s.status_signal.connect(self.updatestatus)
        self.onlinebutton.clicked.connect(self.gonline)
        with open('credentials.txt','r') as g:
            text = g.readlines() 
            self.email = text[0].split(',')[0]
            self.passw = text[0].split(',')[1]
        g.close()
        self.goingonlinelabel.hide()
        self.onlinelabel.hide()
                
    def gonline(self):
        self.movie2 = QMovie('.\\Resources\\goingonline.gif')
        self.goingonlinelabel.setMovie(self.movie2)
        self.movie2.start()
        self.offlinelabel.hide()
        self.goingonlinelabel.show()
        self.hidechrome = self.checkBox.isChecked()
        self.onthread = goonlinethread(self.hidechrome,self.email,self.passw)
        self.onthread.start()
        self.onthread.finished.connect(self.finishOnline)
        
        
    def readandset(self,typee):

        self.textfile = open('textboxesdata.txt','w',encoding='utf-8')
        self.textfile.write('')
        self.textbox.clear()
        typee=self.comboBox.currentText()

        if typee=='Contracts':
            excelfile = openpyxl.load_workbook('.\\Excel Sheets\\Contracts.xlsx')
            cells = ['A','B','C','D','E','F']
            statuscell = 'H'
        elif typee=='Payment Certificates':
            excelfile = openpyxl.load_workbook('.\\Excel Sheets\\Payment Certificates.xlsx')
            cells = ['A','B','C','D','E','F','G']
            statuscell = 'I'
        elif typee=='Variation Requests':
            cells = ['A','B','C','D','E','F','G']
            statuscell = 'I'
            excelfile = openpyxl.load_workbook('.\\Excel Sheets\\Variation Requests.xlsx')
        elif typee=='Variation Orders':
            cells = ['A','B','C','D','E','F']
            statuscell = 'I'
            excelfile = openpyxl.load_workbook('.\\Excel Sheets\\Variation Orders.xlsx')
        else:
            print('There is something wrong with the selection')
            
        sheet = excelfile['Sheet1']
        startindex= 2

        while True:
            startindex+=1
            row = 'A' + str(startindex)
            if type(sheet[row].value)==str:
                pass
            else:
                break
            
        self.textbox.append(f"\t\t\t\t        In Process {startindex-3}\n\n")
        self.start=3

        while self.start!=startindex:
            project_data=[]
            for cellc in cells:
                cell = cellc + str(self.start)
                cellvalue = sheet[cell].value
                project_data.append(cellvalue)
            
            self.textbox.append('\t\t\t\t------------**----------- \n')
            self.textbox.append(f'\t\t\t\t            ({self.start-2}/{startindex-3}) \n')
            self.textfile = open('textboxesdata.txt','a', encoding='utf-8')
            inProcessSh = excelfile['InProcess']
            cell = statuscell+str(self.start)
            status = inProcessSh[cell].value
            
            for value in project_data:
                self.textbox.append(value + '\n')
                self.textfile.write(value+'\n')
            try:
                self.textfile.write('\t\t\t'+status+'\n')
            except TypeError:
                pass
            self.textfile.write('*'+'\n')
            self.start+=1
        self.textbox.verticalScrollBar().setValue(self.textbox.verticalScrollBar().minimum())
        
        excelfile.close()
        self.textfile.close()
        

    def openexcel(self):
        typee=self.currenselection.text()
        if typee=='Contracts':
            QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile('.\\Excel Sheets\\Contracts.xlsx'))
        elif typee=='Payment Certificates':
            QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile('.\\Excel Sheets\\Payment Certificates.xlsx'))
        elif typee=='Variation Requests':
            QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile('.\\Excel Sheets\\Variation Requests.xlsx'))
        elif typee=='Variation Orders':
            QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile('.\\Excel Sheets\\Variation Orders.xlsx'))
        else:
            print('There is something wrong with the selection')  
    
    
    def changeaspect(self):
        typee=self.comboBox.currentText()
        self.currenselection.setText(typee)


    def actionsWindow(self):
        w = Actionswindow()
        win = QDialog()
        w.localview(win)
        n=self.start-3
        print(n)
        w.label.setText(self.currenselection.text())
        if self.currenselection.text()=='Payment Certificates':
                w.label.setGeometry(QtCore.QRect(415, 10, 391, 61))
        if n==0:
            w.emptybox()
        else:
            w.itemshandler(n)
        win.exec()
    
        
    def getupdate(self):
        
        self.notaborted = True
        font2 = QtGui.QFont()
        font2.setFamily("Segoe UI Black")
        font2.setPointSize(10)
        font2.setWeight(50) 
        self.textbox.setFont(font2)
        self.textbox.setText('\t\t\t Please Rest While Crawling For Data\n\t\t\t     This may take up to 3 minutes')
        self.hidechrome = self.checkBox.isChecked()
        self.all = self.updateall.isChecked()
        self.checkBox.hide()
        self.loading.show()
        self.movie.start()
        self.updatestautsbutton.setDisabled(True)
        self.comboBox.setDisabled(True)
        self.timer.show()
        self.timerr = QtCore.QTimer()
        self.counter = 0
        self.timerr.timeout.connect(self.update_counter)
        self.timerr.start(1000)
        self.abortupdatebut.show()
        
        try:
            typee=self.currenselection.text()
            if typee=='Contracts':
                self.workingthread1 = worker1(self.hidechrome,self.email,self.passw)
                self.workingthread1.start()
                self.workingthread1.finished.connect(self.finish)
            elif typee=='Payment Certificates':
                self.workingthread2 = worker2(self.hidechrome,self.email,self.passw)
                self.workingthread2.start()
                self.workingthread2.finished.connect(self.finish)
            elif typee=='Variation Requests':
                self.workingthread3 = worker3(self.hidechrome,self.email,self.passw)
                self.workingthread3.start()
                self.workingthread3.finished.connect(self.finish)
            elif typee=='Variation Orders':
                self.workingthread4 = worker4(self.hidechrome,self.email,self.passw)
                self.workingthread4.start()
                self.workingthread4.finished.connect(self.finish)
            else:
                print('There is something wrong with the selection')
        except:
            self.textbox.setText('Something Went Wrong Try to Run the Update Again')
            self.finish(status=1,aborted=False)
            
        self.s = sender()
        self.s.status_signal.connect(self.updatestatus)
           
    def updatestatus(self,stat):
        print('Update moved to main window')
        self.textbox.append(stat)
    
    def update_counter(self):
        self.counter += 1
        minutes = self.counter // 60
        seconds = self.counter % 60
        time_text = "{:02d}:{:02d}".format(minutes, seconds)
        self.timer.setText(time_text)      
        
    def loginedit(self):
        
        x=Ui_Dialog()
        n=QDialog()
        x.setupUi(n)
        n.setWindowFlags(n.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        n.setWindowFlags(n.windowFlags() | Qt.WindowMinimizeButtonHint)
        n.setWindowIcon(QtGui.QIcon('logo.jpg'))
        n.setFixedSize(772,332)    
            
        with open('credentials.txt','r') as g:
            text = g.readlines() 
            email = text[0].split(',')[0]
            passw = text[0].split(',')[1]
        g.close()
        
        def savecred():
            newemail = x.emailbox.text()
            newpassw = x.passwordbox.text()
            with open('credentials.txt','w') as w :
                w.write(newemail + ',' + newpassw)
            w.close()        
            
            m = QtWidgets.QMessageBox()
            m.setText('\nData Saved\t\t\t\n')
            m.setWindowTitle('Done')
            m.setWindowIcon(QtGui.QIcon('logo.jpg'))
            m.setBaseSize(100,100)
            m.exec()            
        
        x.emailbox.setText(email)
        x.passwordbox.setText(passw)
        x.pushButton.clicked.connect(savecred)
        x.pushButton_2.clicked.connect(n.close)
        
        
        n.exec()
        
    
    def abort(self):
        self.groupBox.setDisabled(True)
        self.notaborted = False
        self.loading.hide()
        self.timer.hide()
        typee=self.currenselection.text()
        self.textbox.setText('\n\n\t\t\t\tAborting Update Please wait')
        self.aborter = killer(typee)
        self.aborter.start()
        self.aborter.finished.connect(self.finish)
        
    def finish(self,aborted,status):
        self.timer.hide()
        self.timerr.stop()
        self.counter=0
        self.textbox.setFont(QtGui.QFont())
        time.sleep(1)
        self.abortupdatebut.hide()
        typee=self.currenselection.text()
        self.loading.hide()
        self.movie.stop()
        self.textbox.clear()
        self.readandset(typee)
        self.updatestautsbutton.setDisabled(False)
        self.comboBox.setDisabled(False)
        self.checkBox.show()
        self.groupBox.setDisabled(False)
        if aborted and status==10:
            m = QtWidgets.QMessageBox()
            m.setText('\nUpdate Aborted By User\t\t\t\n')
            m.setWindowTitle('Aborted')
            m.setWindowIcon(QtGui.QIcon('logo.jpg'))
            m.setFixedWidth(100)
            m.exec()
        elif self.notaborted:
            if status==1:
                self.openexcel()
            elif status==0:
                m = QtWidgets.QMessageBox()
                m.setText('\nNo Internet\t\t\t\n')
                m.setWindowTitle('Connection Error')
                m.setWindowIcon(QtGui.QIcon('logo.jpg'))
                m.setFixedWidth(100)
                m.exec()
        else:
            pass
        
    def finishOnline(self,aborted,status):
        self.movie3 = QMovie('.\\Resources\\online.gif')
        self.onlinelabel.setMovie(self.movie3)
        self.movie3.start()
        self.goingonlinelabel.hide()
        self.offlinelabel.hide()
        self.onlinelabel.show()

if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    container = QtWidgets.QMainWindow()
    windowset = mainwindow()
    windowset.maindisp(container)
    container.show()
    app.exec_()